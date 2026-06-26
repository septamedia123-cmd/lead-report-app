import io
import smtplib
import zipfile
from copy import copy
import pandas as pd
import streamlit as st
import gspread
from email.message import EmailMessage
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

st.set_page_config(
    page_title="LIVMED Report Portal",
    page_icon="📊",
    layout="wide"
)

CC_EMAIL = "erica@livmed.us"
LOGO_FILE = "logo-fixed.png"
GSHEET_ID = st.secrets["GSHEET_ID"]

LOG_HEADERS = [
    "Timestamp",
    "ReportType",
    "TotalRows",
    "PayableLeads",
    "Centers",
    "EmailsSent",
    "Mode",
    "MissingEmails"
]

CENTER_LOG_HEADERS = [
    "Timestamp",
    "ReportType",
    "CenterName",
    "Identifier",
    "TotalRows",
    "PayableLeads",
    "Email",
    "MissingEmail",
    "Mode"
]

ERROR_LOG_HEADERS = [
    "Timestamp",
    "ReportType",
    "ErrorType",
    "Details",
    "Count",
    "Mode"
]

CENTER_PROFILE_HEADERS = [
    "CenterName",
    "Country",
    "Address",
    "Phone",
    "ContactPerson",
    "CommunicationPreference",
    "TeamEmail",
    "Campaign",
    "PaymentSource",
    "PaymentEmail",
    "PaymentDetails",
    "CGMIdentifier",
    "CGMDID",
    "ECPIdentifier",
    "ECPDID",
    "BGMIdentifier",
    "BGMDID",
    "MAIdentifier",
    "MADID",
    "Notes",
    "Active"
]


# =========================
# HELPERS
# =========================
def normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_identifier(value):
    return normalize_text(value)


def find_column(df, candidates):
    lower_map = {str(col).strip().lower(): col for col in df.columns}

    for candidate in candidates:
        c = candidate.lower()
        if c in lower_map:
            return lower_map[c]

    for candidate in candidates:
        c = candidate.lower()
        for col in df.columns:
            if c in str(col).strip().lower():
                return col

    return None


def sanitize_filename(text):
    text = normalize_text(text)
    if not text:
        return "Unknown"
    return "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in text)


def build_email_subject(center_name, identifier, report_name):
    center_name = normalize_text(center_name) or "Center"
    identifier = normalize_text(identifier)
    return f"{report_name} - {center_name} - {identifier}"


def build_email_body(center_name, identifier, total_rows, payable_y, report_name):
    center_name = normalize_text(center_name) or "Team"
    identifier = normalize_text(identifier)

    return f"""Hello {center_name},

Attached is your {report_name.lower()}.

Identifier: {identifier}
Total Records: {total_rows}
Payable Leads: {payable_y}

Please review and let us know if you have any questions.

Best,
Dean
"""


def metric_card(label, value):
    st.markdown(
        f"""
        <div style="
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            border: 1px solid #e5e7eb;
            border-radius: 16px;
            padding: 18px 16px;
            box-shadow: 0 4px 14px rgba(0,0,0,0.05);">
            <div style="font-size: 12px; color: #6b7280; font-weight: 700; text-transform: uppercase; letter-spacing: .04em;">{label}</div>
            <div style="font-size: 28px; color: #111827; font-weight: 800; margin-top: 8px;">{value}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def show_mode_banner(test_mode):
    if test_mode:
        st.markdown(
            """
            <div style="
                background: linear-gradient(90deg, #fff7d6 0%, #fff2b8 100%);
                border: 1px solid #e6c65c;
                padding: 16px 18px;
                border-radius: 12px;
                margin-bottom: 18px;
                font-weight: 700;
                color: #6b5200;
                font-size: 16px;">
                TEST MODE ACTIVE — All emails will be sent only to the test email address.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div style="
                background: linear-gradient(90deg, #ffe0e0 0%, #ffd0d0 100%);
                border: 1px solid #df7b7b;
                padding: 16px 18px;
                border-radius: 12px;
                margin-bottom: 18px;
                font-weight: 800;
                color: #8b0000;
                font-size: 16px;">
                LIVE MODE ACTIVE — Emails will be sent to actual vendor addresses.
            </div>
            """,
            unsafe_allow_html=True,
        )


def show_top_header():
    left, right = st.columns([1, 4])

    with left:
        st.image(LOGO_FILE, width=180)

    with right:
        st.markdown(
            """
            <div style="padding-top: 20px;">
                <div style="font-size: 34px; font-weight: 800; color: #111827;">
                    LIVMED Report Portal
                </div>
                <div style="font-size: 16px; color: #6b7280; margin-top: 4px;">
                    Secure report processing, center profiles, file splitting, and vendor delivery
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<hr style='margin-top: 10px; margin-bottom: 20px;'>", unsafe_allow_html=True)


def go_to(page_name):
    st.session_state.current_page = page_name
    st.rerun()


def open_center_profile(center_name):
    st.session_state.selected_center_name = center_name
    st.session_state.current_page = "profile_detail"
    st.rerun()


def apply_date_filter(df, option):
    if df.empty or "Timestamp" not in df.columns:
        return df

    now = pd.Timestamp.now()

    if option == "Last 7 days":
        cutoff = now - pd.Timedelta(days=7)
        return df[df["Timestamp"] >= cutoff]

    if option == "Last 30 days":
        cutoff = now - pd.Timedelta(days=30)
        return df[df["Timestamp"] >= cutoff]

    return df


# =========================
# GOOGLE SHEETS
# =========================
def get_gsheet_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(
        st.secrets["gcp_service_account"],
        scopes=scopes
    )
    return gspread.authorize(creds)


def get_spreadsheet():
    gc = get_gsheet_client()
    return gc.open_by_key(GSHEET_ID)


def get_or_create_worksheet(title, headers, rows=1000, cols=40):
    spreadsheet = get_spreadsheet()
    try:
        worksheet = spreadsheet.worksheet(title)
    except Exception:
        worksheet = spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)
        worksheet.append_row(headers, value_input_option="USER_ENTERED")

    values = worksheet.get_all_values()
    if not values:
        worksheet.append_row(headers, value_input_option="USER_ENTERED")

    return worksheet


def load_sheet_as_df(title, expected_cols):
    try:
        worksheet = get_or_create_worksheet(title, expected_cols)
        values = worksheet.get_all_values()

        if not values or len(values) < 2:
            return pd.DataFrame(columns=expected_cols)

        headers = values[0]
        rows = values[1:]

        seen = {}
        unique_headers = []
        for h in headers:
            h = str(h).strip()
            if h in seen:
                seen[h] += 1
                unique_headers.append(f"{h}_{seen[h]}")
            else:
                seen[h] = 0
                unique_headers.append(h)

        df = pd.DataFrame(rows, columns=unique_headers)

        rename_map = {}
        for col in df.columns:
            clean = str(col).strip()
            for expected in expected_cols:
                if clean.startswith(expected):
                    rename_map[col] = expected
                    break

        df = df.rename(columns=rename_map)

        for col in expected_cols:
            if col not in df.columns:
                df[col] = ""

        return df[expected_cols]

    except Exception as e:
        st.warning(f"Could not load worksheet '{title}': {e}")
        return pd.DataFrame(columns=expected_cols)


def write_dataframe_to_sheet(title, df, headers):
    worksheet = get_or_create_worksheet(title, headers)

    clean = df.copy()
    for col in headers:
        if col not in clean.columns:
            clean[col] = ""
    clean = clean[headers].fillna("")

    worksheet.clear()
    data = [headers] + clean.astype(str).values.tolist()
    worksheet.update(data, value_input_option="USER_ENTERED")


def log_report_run(report_type, total_rows, payable_leads, centers, emails_sent, mode, missing_emails):
    worksheet = get_or_create_worksheet("logs", LOG_HEADERS)
    row = [
        pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        report_type,
        total_rows,
        payable_leads,
        centers,
        emails_sent,
        mode,
        missing_emails
    ]
    worksheet.append_row(row, value_input_option="USER_ENTERED")


def log_center_runs(report_type, summary_df, mode):
    worksheet = get_or_create_worksheet("center_logs", CENTER_LOG_HEADERS)
    rows = []
    ts = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")

    for _, row in summary_df.iterrows():
        rows.append([
            ts,
            report_type,
            normalize_text(row.get("CenterName", "")),
            normalize_text(row.get("Identifier", "")),
            int(pd.to_numeric(row.get("TotalRows", 0), errors="coerce") or 0),
            int(pd.to_numeric(row.get("PayableY", 0), errors="coerce") or 0),
            normalize_text(row.get("Email", "")),
            1 if normalize_text(row.get("Email", "")) == "" else 0,
            mode
        ])

    if rows:
        worksheet.append_rows(rows, value_input_option="USER_ENTERED")


def log_error_event(report_type, error_type, details, count, mode):
    worksheet = get_or_create_worksheet("error_logs", ERROR_LOG_HEADERS)
    row = [
        pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        report_type,
        error_type,
        details,
        count,
        mode
    ]
    worksheet.append_row(row, value_input_option="USER_ENTERED")


def load_dashboard_logs():
    df = load_sheet_as_df("logs", LOG_HEADERS)
    if "Timestamp" in df.columns:
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
    for col in ["TotalRows", "PayableLeads", "Centers", "EmailsSent", "MissingEmails"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def load_center_logs():
    df = load_sheet_as_df("center_logs", CENTER_LOG_HEADERS)
    if "Timestamp" in df.columns:
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
    for col in ["TotalRows", "PayableLeads", "MissingEmail"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def load_error_logs():
    df = load_sheet_as_df("error_logs", ERROR_LOG_HEADERS)
    if "Timestamp" in df.columns:
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
    if "Count" in df.columns:
        df["Count"] = pd.to_numeric(df["Count"], errors="coerce").fillna(0)
    return df


def load_center_profiles():
    df = load_sheet_as_df("center_profiles", CENTER_PROFILE_HEADERS)
    for col in CENTER_PROFILE_HEADERS:
        if col in df.columns:
            df[col] = df[col].apply(normalize_text)
    return df


def save_center_profiles(df):
    write_dataframe_to_sheet("center_profiles", df, CENTER_PROFILE_HEADERS)


# =========================
# CENTER PROFILE LOOKUPS
# =========================
def build_profile_lookup(profiles_df, id_col_name):
    if profiles_df.empty or id_col_name not in profiles_df.columns:
        return pd.DataFrame(columns=["Identifier", "CenterName", "Email", "ProfileNotes", "ContactPerson"])

    lookup = pd.DataFrame({
        "Identifier": profiles_df[id_col_name].apply(normalize_identifier),
        "CenterName": profiles_df["CenterName"].apply(normalize_text) if "CenterName" in profiles_df.columns else "",
        "Email": profiles_df["TeamEmail"].apply(normalize_text) if "TeamEmail" in profiles_df.columns else "",
        "ProfileNotes": profiles_df["Notes"].apply(normalize_text) if "Notes" in profiles_df.columns else "",
        "ContactPerson": profiles_df["ContactPerson"].apply(normalize_text) if "ContactPerson" in profiles_df.columns else "",
    })

    lookup = lookup[lookup["Identifier"] != ""]
    lookup = lookup.drop_duplicates(subset=["Identifier"])
    return lookup


def merge_with_profile_lookup(base_df, identifier_col, profile_lookup):
    base = base_df.copy()
    base["Identifier_normalized"] = base[identifier_col].apply(normalize_identifier)

    if profile_lookup is None or profile_lookup.empty:
        base["FinalCenterName"] = ""
        base["FinalEmail"] = ""
        base["ProfileNotes"] = ""
        base["ContactPerson"] = ""
        return base

    merged = base.merge(
        profile_lookup,
        left_on="Identifier_normalized",
        right_on="Identifier",
        how="left"
    )

    merged["FinalCenterName"] = merged["CenterName"].fillna("").apply(normalize_text)
    merged["FinalEmail"] = merged["Email"].fillna("").apply(normalize_text)
    merged["ProfileNotes"] = merged["ProfileNotes"].fillna("").apply(normalize_text)
    merged["ContactPerson"] = merged["ContactPerson"].fillna("").apply(normalize_text)

    return merged


# =========================
# FILE / EMAIL HELPERS
# =========================
def send_vendor_emails(vendor_files, sender_email, gmail_app_password, test_mode, test_email, report_name):
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender_email, gmail_app_password)

    sent_count = 0

    for item in vendor_files:
        if not test_mode and not item["Email"]:
            continue

        to_email = test_email.strip() if test_mode else item["Email"]

        msg = EmailMessage()
        msg["Subject"] = build_email_subject(item["CenterName"], item["Identifier"], report_name)
        msg["From"] = sender_email
        msg["To"] = to_email
        msg["CC"] = item["CC"]
        msg.set_content(
            build_email_body(
                item["CenterName"],
                item["Identifier"],
                item["TotalRows"],
                item["PayableY"],
                report_name
            )
        )

        msg.add_attachment(
            item["CSVBytes"],
            maintype="text",
            subtype="csv",
            filename=item["FileName"]
        )

        server.send_message(msg)
        sent_count += 1

    server.quit()
    return sent_count


def get_file_extension(uploaded_file):
    name = uploaded_file.name.lower()
    if "." not in name:
        return ""
    return name.rsplit(".", 1)[1]


def validate_uploaded_file(uploaded_file, expected_type, report_name):
    ext = get_file_extension(uploaded_file)

    if expected_type == "excel" and ext not in ["xlsx", "xls"]:
        st.error(f"{report_name} requires an Excel file (.xlsx or .xls). You uploaded: {uploaded_file.name}")
        st.info("Please upload the original Excel workbook for this report.")
        return False

    if expected_type == "csv" and ext != "csv":
        st.error(f"{report_name} requires a CSV file (.csv). You uploaded: {uploaded_file.name}")
        st.info("Please upload the CSV export for this report.")
        return False

    return True


def render_upload_instructions(report_name, expected_type):
    if expected_type == "excel":
        st.info(
            f"Upload the **{report_name} Excel workbook** in `.xlsx` or `.xls` format. "
            "For CGM and Med Advantage, the workbook should include the required tabs."
        )
    else:
        st.info(f"Upload the **{report_name} CSV file** in `.csv` format.")


def render_excel_tab_requirements(report_name):
    if report_name in ["CGM Report", "Med Advantage Report"]:
        st.caption("Expected workbook tabs: **Detail** and ideally **Conversion Stats**.")


# =========================
# DASHBOARD / UI
# =========================
def dashboard_card(title, subtitle, button_text, page_key):
    st.markdown(
        f"""
        <div style="
            background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
            border: 1px solid #e5e7eb;
            border-radius: 18px;
            padding: 24px;
            min-height: 180px;
            box-shadow: 0 6px 18px rgba(0,0,0,0.05);">
            <div style="font-size: 22px; font-weight: 800; color: #111827;">{title}</div>
            <div style="font-size: 14px; color: #6b7280; margin-top: 8px;">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True
    )
    if st.button(button_text, key=f"btn_{page_key}", width="stretch"):
        go_to(page_key)


def render_dashboard():
    logs_df = load_dashboard_logs()
    center_logs_df = load_center_logs()
    error_logs_df = load_error_logs()

    st.markdown(
        """
        <div style="font-size: 28px; font-weight: 800; color: #111827; margin-bottom: 6px;">
            Dashboard
        </div>
        <div style="font-size: 15px; color: #6b7280; margin-bottom: 24px;">
            Live performance across all report workflows.
        </div>
        """,
        unsafe_allow_html=True
    )

    filter_col1, _ = st.columns([1, 5])
    with filter_col1:
        date_filter = st.selectbox("Date Range", ["All time", "Last 7 days", "Last 30 days"], index=0)

    logs_filtered = apply_date_filter(logs_df, date_filter)
    center_logs_filtered = apply_date_filter(center_logs_df, date_filter)
    error_logs_filtered = apply_date_filter(error_logs_df, date_filter)

    total_runs = int(len(logs_filtered)) if not logs_filtered.empty else 0
    total_emails = int(logs_filtered["EmailsSent"].sum()) if not logs_filtered.empty else 0
    total_payable = int(logs_filtered["PayableLeads"].sum()) if not logs_filtered.empty else 0
    total_leads = int(logs_filtered["TotalRows"].sum()) if not logs_filtered.empty else 0
    total_missing_emails = int(logs_filtered["MissingEmails"].sum()) if not logs_filtered.empty else 0
    conversion_rate = f"{(total_payable / total_leads * 100):.1f}%" if total_leads > 0 else "0.0%"
    last_run = "N/A"
    if not logs_filtered.empty and logs_filtered["Timestamp"].notna().any():
        last_run = logs_filtered["Timestamp"].max().strftime("%Y-%m-%d %H:%M")

    m1, m2, m3, m4, m5, m6 = st.columns(6)
    with m1:
        metric_card("Total Runs", total_runs)
    with m2:
        metric_card("Total Leads", total_leads)
    with m3:
        metric_card("Payable Leads", total_payable)
    with m4:
        metric_card("Conversion Rate", conversion_rate)
    with m5:
        metric_card("Emails Sent", total_emails)
    with m6:
        metric_card("Last Run", last_run)

    st.markdown("### Workflows")
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        dashboard_card("CGM Report", "Upload Excel, split by LeadSource, and send vendor files.", "Open CGM", "cgm")
    with c2:
        dashboard_card("ECP Report", "Upload CSV, calculate payable, split by Sub Id, and send.", "Open ECP", "ecp")
    with c3:
        dashboard_card("Med Advantage", "Upload Excel, split by LeadSource, and send vendor files.", "Open Med Adv", "medadv")
    with c4:
        dashboard_card("Center Profiles", "Manage master center cards, identifiers, contact info, and notes.", "Open Profiles", "profiles")

    if logs_filtered.empty:
        st.info("No dashboard history found yet for this date range. Send at least one report to populate analytics.")
        return

    st.markdown("### Activity Over Time")
    trend_df = logs_filtered.copy()
    trend_df["Date"] = trend_df["Timestamp"].dt.date
    daily = trend_df.groupby("Date")[["TotalRows", "PayableLeads", "EmailsSent"]].sum()
    st.line_chart(daily, width="stretch")

    ch1, ch2 = st.columns(2)
    with ch1:
        st.markdown("### Runs by Report Type")
        runs_by_report = logs_filtered.groupby("ReportType").size()
        st.bar_chart(runs_by_report, width="stretch")
    with ch2:
        st.markdown("### Emails Sent by Report")
        emails_by_report = logs_filtered.groupby("ReportType")["EmailsSent"].sum()
        st.bar_chart(emails_by_report, width="stretch")

    st.markdown("### Top Performing Centers")
    if center_logs_filtered.empty:
        st.info("No center-level history available yet.")
    else:
        top_centers = (
            center_logs_filtered.groupby("CenterName")[["TotalRows", "PayableLeads"]]
            .sum()
            .reset_index()
            .sort_values(by=["PayableLeads", "TotalRows"], ascending=False)
            .head(10)
        )
        st.dataframe(top_centers, width="stretch")

    st.markdown("### Error Tracking")
    e1, e2, e3 = st.columns(3)
    with e1:
        metric_card("Missing Emails", total_missing_emails)
    with e2:
        recent_error_count = int(error_logs_filtered["Count"].sum()) if not error_logs_filtered.empty else 0
        metric_card("Logged Errors", recent_error_count)
    with e3:
        live_runs = int((logs_filtered["Mode"] == "LIVE").sum()) if "Mode" in logs_filtered.columns else 0
        metric_card("Live Runs", live_runs)

    if not error_logs_filtered.empty:
        st.markdown("### Recent Errors")
        recent_errors = error_logs_filtered.sort_values(by="Timestamp", ascending=False).head(10)
        st.dataframe(recent_errors, width="stretch")

    st.markdown("### Recent Activity")
    recent_cols = ["Timestamp", "ReportType", "TotalRows", "PayableLeads", "Centers", "EmailsSent", "Mode", "MissingEmails"]
    recent_df = logs_filtered.sort_values(by="Timestamp", ascending=False).head(10)[recent_cols]
    st.dataframe(recent_df, width="stretch")


def render_center_profiles_page():
    st.markdown(
        """
        <div style="font-size: 28px; font-weight: 800; color: #111827;">
            Center Profiles
        </div>
        <div style="font-size: 15px; color: #6b7280; margin-top: 4px; margin-bottom: 16px;">
            Master directory for all call centers, contacts, notes, and report identifiers.
        </div>
        """,
        unsafe_allow_html=True
    )

    profiles_df = load_center_profiles()

    st.markdown("### Add New Center")

    a1, a2, a3 = st.columns(3)

    with a1:
        new_center_name = st.text_input("Center Name")
        new_country = st.text_input("Country")
        new_contact = st.text_input("Contact Person")
        new_team_email = st.text_input("Team Email")
        new_phone = st.text_input("Phone")

    with a2:
        new_cgm_id = st.text_input("CGM Identifier")
        new_ecp_id = st.text_input("ECP Identifier")
        new_ma_id = st.text_input("MA Identifier")
        new_campaign = st.text_input("Campaign")
        new_comm_pref = st.text_input("Communication Preference")

    with a3:
        new_payment_source = st.text_input("Payment Source")
        new_payment_email = st.text_input("Payment Email")
        new_payment_details = st.text_input("Payment Details")
        new_notes = st.text_area("Notes")
        new_active = st.selectbox("Active", ["Yes", "No"], index=0)

    if st.button("Save New Center", type="primary", width="stretch"):
        if not new_center_name.strip():
            st.error("Center Name is required.")
        else:
            new_row = {col: "" for col in CENTER_PROFILE_HEADERS}
            new_row["CenterName"] = new_center_name
            new_row["Country"] = new_country
            new_row["ContactPerson"] = new_contact
            new_row["TeamEmail"] = new_team_email
            new_row["Phone"] = new_phone
            new_row["CGMIdentifier"] = new_cgm_id
            new_row["ECPIdentifier"] = new_ecp_id
            new_row["MAIdentifier"] = new_ma_id
            new_row["Campaign"] = new_campaign
            new_row["CommunicationPreference"] = new_comm_pref
            new_row["PaymentSource"] = new_payment_source
            new_row["PaymentEmail"] = new_payment_email
            new_row["PaymentDetails"] = new_payment_details
            new_row["Notes"] = new_notes
            new_row["Active"] = new_active

            updated_df = pd.concat([profiles_df, pd.DataFrame([new_row])], ignore_index=True)
            updated_df = updated_df.fillna("")
            updated_df["CenterName"] = updated_df["CenterName"].apply(normalize_text)
            updated_df = updated_df[updated_df["CenterName"] != ""]
            updated_df = updated_df.drop_duplicates(subset=["CenterName"], keep="last")
            save_center_profiles(updated_df)
            st.success("New center saved.")
            st.rerun()

    st.markdown("### Center Directory")

    search_text = st.text_input("Search centers")
    filtered_df = profiles_df.copy()

    if search_text.strip():
        mask = filtered_df.apply(
            lambda col: col.astype(str).str.contains(search_text, case=False, na=False)
        ).any(axis=1)
        filtered_df = filtered_df[mask]

    if filtered_df.empty:
        st.info("No centers found.")
    else:
        for idx, row in filtered_df.iterrows():
            with st.container():
                st.markdown(
                    f"""
                    <div style="
                        background: linear-gradient(180deg, #ffffff 0%, #f8fafc 100%);
                        border: 1px solid #e5e7eb;
                        border-radius: 16px;
                        padding: 18px;
                        margin-bottom: 14px;
                        box-shadow: 0 4px 12px rgba(0,0,0,0.04);">
                        <div style="font-size: 20px; font-weight: 800; color: #111827;">
                            {normalize_text(row.get('CenterName', ''))}
                        </div>
                        <div style="font-size: 13px; color: #6b7280; margin-top: 4px;">
                            {normalize_text(row.get('Country', ''))} • {normalize_text(row.get('ContactPerson', ''))}
                        </div>
                        <div style="margin-top: 8px; font-size: 13px;">
                            📧 {normalize_text(row.get('TeamEmail', ''))}
                        </div>
                        <div style="margin-top: 8px; font-size: 12px; color: #374151;">
                            CGM: {normalize_text(row.get('CGMIdentifier', ''))} |
                            ECP: {normalize_text(row.get('ECPIdentifier', ''))} |
                            MA: {normalize_text(row.get('MAIdentifier', ''))}
                        </div>
                        <div style="margin-top: 8px; font-size: 12px; color: #6b7280;">
                            {normalize_text(row.get('Notes', ''))}
                        </div>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

                c1, c2, c3 = st.columns(3)
                with c1:
                    if st.button("Open Profile", key=f"open_center_{idx}", width="stretch"):
                        open_center_profile(normalize_text(row.get("CenterName", "")))
                with c2:
                    if st.button("Edit", key=f"edit_center_{idx}", width="stretch"):
                        st.session_state["edit_center_idx"] = idx
                        st.rerun()
                with c3:
                    if st.button("Delete", key=f"delete_center_{idx}", width="stretch"):
                        updated_df = profiles_df.drop(index=idx).reset_index(drop=True)
                        save_center_profiles(updated_df)
                        st.warning("Center deleted.")
                        st.rerun()

    if "edit_center_idx" in st.session_state:
        edit_idx = st.session_state["edit_center_idx"]
        edit_row = profiles_df.loc[edit_idx]

        st.markdown("### Edit Center")

        e1, e2, e3 = st.columns(3)

        with e1:
            edit_center_name = st.text_input("Edit Center Name", value=edit_row.get("CenterName", ""), key="edit_center_name")
            edit_country = st.text_input("Edit Country", value=edit_row.get("Country", ""), key="edit_country")
            edit_contact = st.text_input("Edit Contact Person", value=edit_row.get("ContactPerson", ""), key="edit_contact")
            edit_team_email = st.text_input("Edit Team Email", value=edit_row.get("TeamEmail", ""), key="edit_team_email")
            edit_phone = st.text_input("Edit Phone", value=edit_row.get("Phone", ""), key="edit_phone")

        with e2:
            edit_cgm_id = st.text_input("Edit CGM Identifier", value=edit_row.get("CGMIdentifier", ""), key="edit_cgm")
            edit_ecp_id = st.text_input("Edit ECP Identifier", value=edit_row.get("ECPIdentifier", ""), key="edit_ecp")
            edit_ma_id = st.text_input("Edit MA Identifier", value=edit_row.get("MAIdentifier", ""), key="edit_ma")
            edit_campaign = st.text_input("Edit Campaign", value=edit_row.get("Campaign", ""), key="edit_campaign")
            edit_comm_pref = st.text_input("Edit Communication Preference", value=edit_row.get("CommunicationPreference", ""), key="edit_comm")

        with e3:
            edit_payment_source = st.text_input("Edit Payment Source", value=edit_row.get("PaymentSource", ""), key="edit_pay_source")
            edit_payment_email = st.text_input("Edit Payment Email", value=edit_row.get("PaymentEmail", ""), key="edit_pay_email")
            edit_payment_details = st.text_input("Edit Payment Details", value=edit_row.get("PaymentDetails", ""), key="edit_pay_details")
            edit_notes = st.text_area("Edit Notes", value=edit_row.get("Notes", ""), key="edit_notes")
            edit_active = st.selectbox("Edit Active", ["Yes", "No"], index=0 if edit_row.get("Active", "Yes") == "Yes" else 1, key="edit_active")

        s1, s2 = st.columns(2)
        with s1:
            if st.button("Save Center Changes", width="stretch"):
                for col in CENTER_PROFILE_HEADERS:
                    if col not in profiles_df.columns:
                        profiles_df[col] = ""

                profiles_df.at[edit_idx, "CenterName"] = edit_center_name
                profiles_df.at[edit_idx, "Country"] = edit_country
                profiles_df.at[edit_idx, "ContactPerson"] = edit_contact
                profiles_df.at[edit_idx, "TeamEmail"] = edit_team_email
                profiles_df.at[edit_idx, "Phone"] = edit_phone
                profiles_df.at[edit_idx, "CGMIdentifier"] = edit_cgm_id
                profiles_df.at[edit_idx, "ECPIdentifier"] = edit_ecp_id
                profiles_df.at[edit_idx, "MAIdentifier"] = edit_ma_id
                profiles_df.at[edit_idx, "Campaign"] = edit_campaign
                profiles_df.at[edit_idx, "CommunicationPreference"] = edit_comm_pref
                profiles_df.at[edit_idx, "PaymentSource"] = edit_payment_source
                profiles_df.at[edit_idx, "PaymentEmail"] = edit_payment_email
                profiles_df.at[edit_idx, "PaymentDetails"] = edit_payment_details
                profiles_df.at[edit_idx, "Notes"] = edit_notes
                profiles_df.at[edit_idx, "Active"] = edit_active

                save_center_profiles(profiles_df)
                st.success("Center updated.")
                del st.session_state["edit_center_idx"]
                st.rerun()

        with s2:
            if st.button("Cancel Edit", width="stretch"):
                del st.session_state["edit_center_idx"]
                st.rerun()

    with st.expander("Advanced Table Editor"):
        edited_df = st.data_editor(profiles_df, width="stretch", num_rows="dynamic", key="profiles_table_editor")
        if st.button("Save Table Changes", width="stretch"):
            save_center_profiles(edited_df)
            st.success("Table changes saved.")
            st.rerun()


def render_center_profile_detail_page():
    profiles_df = load_center_profiles()
    center_logs_df = load_center_logs()
    error_logs_df = load_error_logs()

    selected_center = normalize_text(st.session_state.get("selected_center_name", ""))

    if not selected_center:
        st.warning("No center selected.")
        if st.button("Back to Center Profiles", width="stretch"):
            go_to("profiles")
        return

    match_df = profiles_df[profiles_df["CenterName"].apply(normalize_text) == selected_center]
    if match_df.empty:
        st.warning("That center could not be found.")
        if st.button("Back to Center Profiles", width="stretch"):
            go_to("profiles")
        return

    row = match_df.iloc[0]

    back_col, title_col = st.columns([1, 6])
    with back_col:
        if st.button("← Back", width="stretch"):
            go_to("profiles")
    with title_col:
        st.markdown(
            f"""
            <div style="font-size: 30px; font-weight: 800; color: #111827;">
                {selected_center}
            </div>
            <div style="font-size: 15px; color: #6b7280; margin-top: 4px;">
                {normalize_text(row.get('Country', ''))} • {normalize_text(row.get('Active', ''))}
            </div>
            """,
            unsafe_allow_html=True
        )

    i1, i2 = st.columns(2)

    with i1:
        st.markdown("### Basic Info")
        basic_info = pd.DataFrame({
            "Field": ["Center Name", "Country", "Address", "Phone", "Contact Person", "Communication Preference", "Team Email", "Campaign", "Active"],
            "Value": [
                row.get("CenterName", ""),
                row.get("Country", ""),
                row.get("Address", ""),
                row.get("Phone", ""),
                row.get("ContactPerson", ""),
                row.get("CommunicationPreference", ""),
                row.get("TeamEmail", ""),
                row.get("Campaign", ""),
                row.get("Active", "")
            ]
        })
        st.dataframe(basic_info, width="stretch", hide_index=True)

    with i2:
        st.markdown("### Report Identifiers")
        ids_df = pd.DataFrame({
            "Report": ["CGM", "ECP", "BGM", "Med Advantage"],
            "Identifier": [
                row.get("CGMIdentifier", ""),
                row.get("ECPIdentifier", ""),
                row.get("BGMIdentifier", ""),
                row.get("MAIdentifier", "")
            ],
            "DID": [
                row.get("CGMDID", ""),
                row.get("ECPDID", ""),
                row.get("BGMDID", ""),
                row.get("MADID", "")
            ]
        })
        st.dataframe(ids_df, width="stretch", hide_index=True)

    p1, p2 = st.columns(2)

    with p1:
        st.markdown("### Payment / Operations")
        payment_df = pd.DataFrame({
            "Field": ["Payment Source", "Payment Email", "Payment Details"],
            "Value": [
                row.get("PaymentSource", ""),
                row.get("PaymentEmail", ""),
                row.get("PaymentDetails", "")
            ]
        })
        st.dataframe(payment_df, width="stretch", hide_index=True)

    with p2:
        st.markdown("### Notes")
        st.text_area("Notes", value=row.get("Notes", ""), height=180, disabled=True, key="profile_notes_view")

    center_history = center_logs_df[center_logs_df["CenterName"].apply(normalize_text) == selected_center].copy()
    center_errors = error_logs_df[
        error_logs_df["Details"].astype(str).str.contains(selected_center, case=False, na=False)
    ].copy()

    total_leads = int(center_history["TotalRows"].sum()) if not center_history.empty else 0
    total_payable = int(center_history["PayableLeads"].sum()) if not center_history.empty else 0
    total_emails = int((center_history["MissingEmail"] == 0).sum()) if not center_history.empty else 0
    missing_email_issues = int(center_history["MissingEmail"].sum()) if not center_history.empty else 0
    conversion_rate = f"{(total_payable / total_leads * 100):.1f}%" if total_leads > 0 else "0.0%"
    last_run = "N/A"
    if not center_history.empty and center_history["Timestamp"].notna().any():
        last_run = center_history["Timestamp"].max().strftime("%Y-%m-%d %H:%M")

    st.markdown("### Performance Snapshot")
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    with m1:
        metric_card("Total Leads", total_leads)
    with m2:
        metric_card("Payable Leads", total_payable)
    with m3:
        metric_card("Conversion Rate", conversion_rate)
    with m4:
        metric_card("Emails Sent", total_emails)
    with m5:
        metric_card("Last Run", last_run)
    with m6:
        metric_card("Missing Email Issues", missing_email_issues)

    st.markdown("### Alerts")
    alerts = []

    if normalize_text(row.get("TeamEmail", "")) == "":
        alerts.append("Missing Team Email.")
    if normalize_text(row.get("CGMIdentifier", "")) == "" and normalize_text(row.get("ECPIdentifier", "")) == "" and normalize_text(row.get("MAIdentifier", "")) == "":
        alerts.append("No active report identifiers assigned.")
    if center_history.empty:
        alerts.append("No report activity logged yet for this center.")
    else:
        most_recent = center_history["Timestamp"].max()
        if pd.notna(most_recent):
            days_since = (pd.Timestamp.now() - most_recent).days
            if days_since > 30:
                alerts.append("No report activity in the last 30 days.")
    if missing_email_issues > 0:
        alerts.append(f"{missing_email_issues} run(s) logged with missing email status.")

    if alerts:
        for alert in alerts:
            st.warning(alert)
    else:
        st.success("No active alerts for this center.")

    st.markdown("### Activity History")
    if center_history.empty:
        st.info("No activity history found for this center.")
    else:
        history_cols = ["Timestamp", "ReportType", "Identifier", "TotalRows", "PayableLeads", "Mode", "MissingEmail"]
        st.dataframe(
            center_history.sort_values(by="Timestamp", ascending=False)[history_cols].head(15),
            width="stretch"
        )

    if not center_errors.empty:
        st.markdown("### Error History")
        err_cols = ["Timestamp", "ReportType", "ErrorType", "Details", "Count", "Mode"]
        st.dataframe(center_errors.sort_values(by="Timestamp", ascending=False)[err_cols].head(10), width="stretch")


# =========================
# REPORT PAGE ENGINE V2
# =========================
REPORT_CONFIGS = {
    "cgm": {
        "report_key": "cgm",
        "report_name": "CGM Report",
        "dashboard_title": "CGM Report",
        "dashboard_subtitle": "Upload Excel, split by LeadSource, review, approve, and send center workbooks.",
        "button_text": "Open CGM",
        "identifier_label": "LeadSource",
        "profile_identifier_col": "CGMIdentifier",
        "file_type": "excel",
        "payable_rule": "existing_payable",
        "remove_columns": ["PaidAmount", "Paid Amount", "DiabeticOnMedicare", "Diabetic On Medicare", "AssignedTo", "Assigned To"],
        "remove_column_indexes": [],
    },
    "ecp": {
        "report_key": "ecp",
        "report_name": "ECP Report",
        "dashboard_title": "ECP Report",
        "dashboard_subtitle": "Upload CSV, calculate payable from duration, review, approve, and send center workbooks.",
        "button_text": "Open ECP",
        "identifier_label": "Sub Id",
        "profile_identifier_col": "ECPIdentifier",
        "file_type": "csv",
        "payable_rule": "duration_column_k_over_630",
        "remove_columns": [],
        "remove_column_indexes": [4],  # Column E contains LivMed name and is hidden from final center reports.
    },
    "medadv": {
        "report_key": "medadv",
        "report_name": "Med Advantage Report",
        "dashboard_title": "Med Advantage",
        "dashboard_subtitle": "Upload Excel, split by LeadSource, review, approve, and send center workbooks.",
        "button_text": "Open Med Adv",
        "identifier_label": "LeadSource",
        "profile_identifier_col": "MAIdentifier",
        "file_type": "excel",
        "payable_rule": "existing_payable",
        "remove_columns": ["PaidAmount", "Paid Amount", "DiabeticOnMedicare", "Diabetic On Medicare", "AssignedTo", "Assigned To"],
        "remove_column_indexes": [],
    },
}

INTERNAL_MERGE_COLUMNS = [
    "Identifier_normalized", "Identifier", "CenterName", "Email", "FinalCenterName",
    "FinalEmail", "ProfileNotes", "ContactPerson"
]


def excel_sheet_name(value):
    name = sanitize_filename(value).replace("_", " ").strip() or "Sheet"
    for bad in [":", "\\", "/", "?", "*", "[", "]"]:
        name = name.replace(bad, "")
    return name[:31] or "Sheet"


def apply_excel_widths(ws):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max_len + 3, 45)


def style_header_row(ws, row_num=1):
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def write_df_to_sheet(wb, sheet_name, df, title=None, freeze=True):
    ws = wb.create_sheet(excel_sheet_name(sheet_name))
    start_row = 1
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(size=16, bold=True)
        start_row = 3

    safe_df = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    if safe_df.empty:
        ws.cell(start_row, 1, "No data available")
        apply_excel_widths(ws)
        return ws

    for c_idx, col in enumerate(safe_df.columns, 1):
        ws.cell(start_row, c_idx, str(col))
    style_header_row(ws, start_row)

    for r_idx, row in enumerate(safe_df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = ""
            ws.cell(r_idx, c_idx, value)

    if freeze:
        ws.freeze_panes = f"A{start_row + 1}"
    apply_excel_widths(ws)
    return ws


def df_to_excel_bytes(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def payable_counts(df):
    col = find_column(df, ["Payable"])
    if not col or col not in df.columns:
        return 0, 0
    s = df[col].astype(str).str.strip().str.upper()
    return int((s == "Y").sum()), int((s == "N").sum())


def prepare_uploaded_detail(uploaded_file, config):
    file_type = config["file_type"]
    if file_type == "excel":
        xls = pd.ExcelFile(uploaded_file)
        detail_sheet = None
        for sheet in xls.sheet_names:
            if "detail" in sheet.strip().lower():
                detail_sheet = sheet
                break
        if detail_sheet is None:
            raise ValueError("Could not find a Detail sheet in this workbook.")
        detail_df = pd.read_excel(uploaded_file, sheet_name=detail_sheet)
    else:
        detail_df = pd.read_csv(uploaded_file)

    detail_df.columns = [str(c).strip() for c in detail_df.columns]
    return detail_df


def apply_payable_rule(detail_df, config):
    df = detail_df.copy()
    rule = config.get("payable_rule", "existing_payable")

    if rule == "duration_column_k_over_630":
        if len(df.columns) <= 10:
            raise ValueError("ECP requires Column K to exist because Column K is the Duration field.")
        duration_col = df.columns[10]  # Column K, zero-based index 10.
        duration_values = pd.to_numeric(df[duration_col], errors="coerce").fillna(0)
        df["Payable"] = duration_values.apply(lambda x: "Y" if x > 630 else "N")
        df["Payable Audit Reason"] = duration_values.apply(
            lambda x: f"Payable: Column K duration {x:g} > 630 seconds" if x > 630
            else f"Not payable: Column K duration {x:g} <= 630 seconds"
        )
        df["Payable Duration Seconds"] = duration_values
        return df

    payable_col = find_column(df, ["Payable"])
    if payable_col and payable_col in df.columns:
        df["Payable Audit Reason"] = df[payable_col].astype(str).apply(
            lambda x: "Used Payable value from uploaded report: " + normalize_text(x)
        )
    else:
        df["Payable"] = ""
        df["Payable Audit Reason"] = "No Payable column found in uploaded report"
    return df


def remove_final_report_columns(group, config):
    cols_to_remove = INTERNAL_MERGE_COLUMNS + config.get("remove_columns", [])
    export_df = group.drop(columns=cols_to_remove, errors="ignore").copy()

    remove_indexes = sorted(config.get("remove_column_indexes", []), reverse=True)
    for idx in remove_indexes:
        if 0 <= idx < len(export_df.columns):
            export_df = export_df.drop(columns=[export_df.columns[idx]])
    return export_df


def build_disposition_summary(df):
    disposition_col = find_column(df, ["Disposition"])
    if disposition_col and disposition_col in df.columns:
        return (
            df.groupby(["Identifier_normalized", disposition_col])
            .size()
            .reset_index(name="Count")
            .rename(columns={"Identifier_normalized": "Identifier"})
        )
    return pd.DataFrame()


def build_duplicate_report(df):
    candidates = ["CallId", "Call ID", "CallShaperLeadId", "Lead ID", "FromNumber", "Phone", "Phone Number"]
    dup_frames = []
    for candidate in candidates:
        col = find_column(df, [candidate])
        if col and col in df.columns:
            temp = df[df[col].astype(str).str.strip() != ""].copy()
            dups = temp[temp.duplicated(subset=[col], keep=False)].copy()
            if not dups.empty:
                dups.insert(0, "Duplicate Check Field", col)
                dup_frames.append(dups)
            break
    if dup_frames:
        return pd.concat(dup_frames, ignore_index=True)
    return pd.DataFrame()


def build_exception_report(merged_df, summary_df):
    rows = []
    if not summary_df.empty:
        for _, row in summary_df.iterrows():
            ident = normalize_text(row.get("Identifier", ""))
            center = normalize_text(row.get("CenterName", ""))
            email = normalize_text(row.get("Email", ""))
            if not center:
                rows.append({"Issue": "Unknown identifier", "Identifier": ident, "CenterName": center, "Email": email, "Details": "Identifier is not mapped to a Center Profile."})
            if not email:
                rows.append({"Issue": "Missing email", "Identifier": ident, "CenterName": center, "Email": email, "Details": "Center has no TeamEmail in Center Profiles."})

    if "Identifier_normalized" in merged_df.columns:
        missing_identifier = merged_df[merged_df["Identifier_normalized"].astype(str).str.strip() == ""]
        if not missing_identifier.empty:
            rows.append({"Issue": "Blank identifier", "Identifier": "", "CenterName": "", "Email": "", "Details": f"{len(missing_identifier)} row(s) have a blank identifier."})

    return pd.DataFrame(rows, columns=["Issue", "Identifier", "CenterName", "Email", "Details"])


def build_email_preview_df(vendor_files, test_mode, test_email):
    rows = []
    for item in vendor_files:
        rows.append({
            "CenterName": item["CenterName"],
            "Identifier": item["Identifier"],
            "ToEmail": test_email.strip() if test_mode else item["Email"],
            "ActualEmail": item["Email"],
            "CC": item["CC"],
            "File": item["FileName"],
            "TotalRows": item["TotalRows"],
            "PayableY": item["PayableY"],
            "Status": "READY" if (test_mode and test_email.strip()) or item["Email"] else "MISSING EMAIL"
        })
    return pd.DataFrame(rows)


def build_center_workbook(center_name, identifier, report_name, export_df, center_history_df, disposition_df=None):
    wb = Workbook()
    wb.remove(wb.active)

    current_payable, current_not_payable = payable_counts(export_df)
    total_rows = len(export_df)
    conversion_rate = current_payable / total_rows if total_rows else 0

    summary = pd.DataFrame([
        {"Metric": "Report", "Value": report_name},
        {"Metric": "Center", "Value": center_name},
        {"Metric": "Identifier", "Value": identifier},
        {"Metric": "Total Leads", "Value": total_rows},
        {"Metric": "Payable Leads", "Value": current_payable},
        {"Metric": "Not Payable", "Value": current_not_payable},
        {"Metric": "Payable %", "Value": f"{conversion_rate:.1%}"},
    ])
    write_df_to_sheet(wb, "Summary", summary, title=f"{center_name or 'Center'} Performance Summary")

    write_df_to_sheet(wb, "Current Leads", export_df, title=f"{report_name} - Current Leads")

    history = center_history_df.copy() if isinstance(center_history_df, pd.DataFrame) else pd.DataFrame()
    if not history.empty:
        history = history.sort_values(by="Timestamp")
        history["RunDate"] = pd.to_datetime(history["Timestamp"], errors="coerce").dt.strftime("%Y-%m-%d")
        history["PayableRate"] = history.apply(
            lambda row: (float(row.get("PayableLeads", 0)) / float(row.get("TotalRows", 0))) if float(row.get("TotalRows", 0) or 0) else 0,
            axis=1
        )
        monthly = history[["RunDate", "ReportType", "TotalRows", "PayableLeads", "PayableRate"]].copy()
    else:
        monthly = pd.DataFrame([{
            "RunDate": pd.Timestamp.now().strftime("%Y-%m-%d"),
            "ReportType": report_name,
            "TotalRows": total_rows,
            "PayableLeads": current_payable,
            "PayableRate": conversion_rate,
        }])

    ws_monthly = write_df_to_sheet(wb, "Monthly Performance", monthly, title="Performance Over Time")
    for row in range(4, ws_monthly.max_row + 1):
        try:
            ws_monthly.cell(row, 5).number_format = "0.00%"
        except Exception:
            pass

    if disposition_df is not None and not disposition_df.empty:
        write_df_to_sheet(wb, "Disposition Summary", disposition_df, title="Disposition Breakdown")

    charts = wb.create_sheet("Charts")
    charts["A1"] = f"{center_name or 'Center'} Charts"
    charts["A1"].font = Font(size=16, bold=True)
    charts["A3"] = "RunDate"
    charts["B3"] = "TotalRows"
    charts["C3"] = "PayableLeads"
    charts["D3"] = "PayableRate"
    style_header_row(charts, 3)

    for idx, row in enumerate(monthly.itertuples(index=False), 4):
        charts.cell(idx, 1, getattr(row, "RunDate", ""))
        charts.cell(idx, 2, float(getattr(row, "TotalRows", 0) or 0))
        charts.cell(idx, 3, float(getattr(row, "PayableLeads", 0) or 0))
        charts.cell(idx, 4, float(getattr(row, "PayableRate", 0) or 0))
        charts.cell(idx, 4).number_format = "0.00%"

    last = max(4, charts.max_row)

    total_chart = LineChart()
    total_chart.title = "Total Leads Over Time"
    total_chart.y_axis.title = "Leads"
    total_chart.x_axis.title = "Run Date"
    total_chart.add_data(Reference(charts, min_col=2, min_row=3, max_row=last), titles_from_data=True)
    total_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=last))
    total_chart.height = 8
    total_chart.width = 14
    charts.add_chart(total_chart, "A8")

    payable_chart = LineChart()
    payable_chart.title = "Payable Leads Over Time"
    payable_chart.y_axis.title = "Payable Leads"
    payable_chart.x_axis.title = "Run Date"
    payable_chart.add_data(Reference(charts, min_col=3, min_row=3, max_row=last), titles_from_data=True)
    payable_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=last))
    payable_chart.height = 8
    payable_chart.width = 14
    charts.add_chart(payable_chart, "I8")

    rate_chart = LineChart()
    rate_chart.title = "Payable Rate Over Time"
    rate_chart.y_axis.title = "Payable %"
    rate_chart.x_axis.title = "Run Date"
    rate_chart.add_data(Reference(charts, min_col=4, min_row=3, max_row=last), titles_from_data=True)
    rate_chart.set_categories(Reference(charts, min_col=1, min_row=4, max_row=last))
    rate_chart.height = 8
    rate_chart.width = 14
    charts.add_chart(rate_chart, "A24")

    if disposition_df is not None and not disposition_df.empty and "Count" in disposition_df.columns:
        disp_start = 3
        charts["F3"] = "Disposition"
        charts["G3"] = "Count"
        style_header_row(charts, 3)
        disp_col = [c for c in disposition_df.columns if c not in ["Identifier", "Count"]]
        label_col = disp_col[0] if disp_col else disposition_df.columns[0]
        for r_idx, (_, disp_row) in enumerate(disposition_df.iterrows(), 4):
            charts.cell(r_idx, 6, normalize_text(disp_row.get(label_col, "")))
            charts.cell(r_idx, 7, int(disp_row.get("Count", 0) or 0))
        pie_last = max(4, 3 + len(disposition_df))
        pie = PieChart()
        pie.title = "Disposition Breakdown"
        pie.add_data(Reference(charts, min_col=7, min_row=3, max_row=pie_last), titles_from_data=True)
        pie.set_categories(Reference(charts, min_col=6, min_row=4, max_row=pie_last))
        pie.height = 8
        pie.width = 14
        charts.add_chart(pie, "I24")

    apply_excel_widths(charts)
    return wb


def build_admin_review_workbook(config, raw_df, merged_df, summary_df, email_preview_df, missing_email_df, unknown_identifier_df, duplicate_df, disposition_summary, exception_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = f"{config['report_name']} Admin Review"
    ws["A1"].font = Font(size=18, bold=True)

    total_rows = int(summary_df["TotalRows"].sum()) if not summary_df.empty else 0
    total_payable = int(summary_df["PayableY"].sum()) if not summary_df.empty else 0
    total_centers = len(summary_df)
    missing_count = len(missing_email_df) if isinstance(missing_email_df, pd.DataFrame) else 0
    unknown_count = len(unknown_identifier_df) if isinstance(unknown_identifier_df, pd.DataFrame) else 0
    duplicate_count = len(duplicate_df) if isinstance(duplicate_df, pd.DataFrame) else 0

    dashboard_rows = [
        ("Report", config["report_name"]),
        ("Generated At", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Centers", total_centers),
        ("Total Leads", total_rows),
        ("Payable Leads", total_payable),
        ("Payable %", f"{(total_payable / total_rows):.1%}" if total_rows else "0.0%"),
        ("Missing Email Centers", missing_count),
        ("Unknown Identifier Rows", unknown_count),
        ("Duplicate Rows Flagged", duplicate_count),
        ("Status", "REVIEW REQUIRED BEFORE SENDING"),
    ]
    for r, (label, value) in enumerate(dashboard_rows, 3):
        ws.cell(r, 1, label).font = Font(bold=True)
        ws.cell(r, 2, value)
    apply_excel_widths(ws)

    write_df_to_sheet(wb, "Center Summary", summary_df, title="Center Summary")
    write_df_to_sheet(wb, "Email Queue", email_preview_df, title="Email Queue Preview")
    write_df_to_sheet(wb, "Missing Emails", missing_email_df, title="Missing Emails")
    write_df_to_sheet(wb, "Unknown Identifiers", unknown_identifier_df, title="Unknown Identifiers")
    write_df_to_sheet(wb, "Duplicate Leads", duplicate_df, title="Duplicate Leads")
    write_df_to_sheet(wb, "Payable Audit", merged_df[[c for c in merged_df.columns if c in ["Identifier_normalized", "FinalCenterName", "Payable", "Payable Audit Reason", "Payable Duration Seconds"]]], title="Payable Audit")
    if not disposition_summary.empty:
        write_df_to_sheet(wb, "Disposition Summary", disposition_summary, title="Disposition Summary")
    write_df_to_sheet(wb, "Exceptions", exception_df, title="Exception Report")
    write_df_to_sheet(wb, "Raw Import", raw_df, title="Raw Uploaded Detail")

    # Add a sheet per center for admin review.
    for identifier_value, group in merged_df.groupby("Identifier_normalized"):
        center_name = normalize_text(group["FinalCenterName"].iloc[0]) if "FinalCenterName" in group.columns else ""
        label = center_name or identifier_value or "Unknown Center"
        write_df_to_sheet(wb, f"{label[:20]}", group, title=f"{label} Detail")

    return wb


def process_report_package(uploaded_file, config, test_mode, test_email):
    raw_df = prepare_uploaded_detail(uploaded_file, config)
    detail_df = apply_payable_rule(raw_df, config)

    id_col = find_column(detail_df, [config["identifier_label"], config["identifier_label"].replace(" ", "")])
    if id_col is None:
        raise ValueError(f"Could not find {config['identifier_label']} in the uploaded file.")

    paidamount_col = find_column(detail_df, ["PaidAmount", "Paid Amount"])
    profiles_df = load_center_profiles()
    profile_lookup = build_profile_lookup(profiles_df, config["profile_identifier_col"])
    merged_df = merge_with_profile_lookup(detail_df, id_col, profile_lookup)

    summary_rows = []
    vendor_files = []
    center_logs_df = load_center_logs()
    disposition_summary = build_disposition_summary(merged_df)

    for identifier_value, group in merged_df.groupby("Identifier_normalized", dropna=False):
        center_name = normalize_text(group["FinalCenterName"].iloc[0]) if "FinalCenterName" in group.columns else ""
        email = normalize_text(group["FinalEmail"].iloc[0]) if "FinalEmail" in group.columns else ""
        payable_y, payable_n = payable_counts(group)
        total_paid = 0.0
        if paidamount_col and paidamount_col in group.columns:
            total_paid = float(pd.to_numeric(group[paidamount_col], errors="coerce").fillna(0).sum())

        summary_rows.append({
            "Identifier": identifier_value,
            "CenterName": center_name,
            "Email": email,
            "CC": CC_EMAIL,
            "TotalRows": len(group),
            "PayableY": payable_y,
            "PayableN": payable_n,
            "TotalPaidAmount": total_paid,
            "ReadyToSend": "Yes" if email else "No"
        })

    summary_df = pd.DataFrame(summary_rows)
    if not summary_df.empty:
        summary_df = summary_df.sort_values(by=["CenterName", "Identifier"], na_position="last")

    missing_email_df = summary_df[summary_df["Email"].astype(str).str.strip() == ""].copy() if not summary_df.empty else pd.DataFrame()
    unknown_identifier_df = merged_df[merged_df["FinalCenterName"].astype(str).str.strip() == ""].copy() if "FinalCenterName" in merged_df.columns else pd.DataFrame()
    duplicate_df = build_duplicate_report(merged_df)
    exception_df = build_exception_report(merged_df, summary_df)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for identifier_value, group in merged_df.groupby("Identifier_normalized", dropna=False):
            center_name = normalize_text(group["FinalCenterName"].iloc[0]) if "FinalCenterName" in group.columns else ""
            email = normalize_text(group["FinalEmail"].iloc[0]) if "FinalEmail" in group.columns else ""
            safe_identifier = sanitize_filename(identifier_value)
            safe_center = sanitize_filename(center_name) if center_name else "UnknownCenter"
            export_group = remove_final_report_columns(group, config)

            center_history = pd.DataFrame()
            if not center_logs_df.empty:
                center_history = center_logs_df[
                    (center_logs_df["Identifier"].astype(str) == str(identifier_value))
                    & (center_logs_df["ReportType"].astype(str) == config["report_name"])
                ].copy()

            center_disposition = pd.DataFrame()
            if not disposition_summary.empty:
                center_disposition = disposition_summary[disposition_summary["Identifier"].astype(str) == str(identifier_value)].copy()

            wb = build_center_workbook(
                center_name=center_name or safe_identifier,
                identifier=identifier_value,
                report_name=config["report_name"],
                export_df=export_group,
                center_history_df=center_history,
                disposition_df=center_disposition,
            )
            workbook_bytes = df_to_excel_bytes(wb)
            workbook_name = f"{safe_center}__{safe_identifier}.xlsx"
            zip_file.writestr(workbook_name, workbook_bytes)

            vendor_files.append({
                "Identifier": identifier_value,
                "CenterName": center_name,
                "Email": email,
                "CC": CC_EMAIL,
                "FileName": workbook_name,
                "FileBytes": workbook_bytes,
                "TotalRows": len(group),
                "PayableY": payable_counts(group)[0],
            })

        zip_file.writestr("center_summary.csv", summary_df.to_csv(index=False).encode("utf-8"))
        if not disposition_summary.empty:
            zip_file.writestr("disposition_summary.csv", disposition_summary.to_csv(index=False).encode("utf-8"))

    zip_buffer.seek(0)
    email_preview_df = build_email_preview_df(vendor_files, test_mode, test_email)
    admin_wb = build_admin_review_workbook(
        config, raw_df, merged_df, summary_df, email_preview_df,
        missing_email_df, unknown_identifier_df, duplicate_df, disposition_summary, exception_df
    )
    admin_workbook_bytes = df_to_excel_bytes(admin_wb)

    total_centers = len(summary_df)
    total_rows = int(summary_df["TotalRows"].sum()) if not summary_df.empty else 0
    total_payable = int(summary_df["PayableY"].sum()) if not summary_df.empty else 0
    ready_count = int((summary_df["ReadyToSend"] == "Yes").sum()) if not summary_df.empty else 0

    return {
        "config": config,
        "raw_df": raw_df,
        "merged_df": merged_df,
        "summary_df": summary_df,
        "missing_email_df": missing_email_df,
        "unknown_identifier_df": unknown_identifier_df,
        "duplicate_df": duplicate_df,
        "disposition_summary": disposition_summary,
        "exception_df": exception_df,
        "email_preview_df": email_preview_df,
        "vendor_files": vendor_files,
        "zip_bytes": zip_buffer.getvalue(),
        "admin_workbook_bytes": admin_workbook_bytes,
        "total_centers": total_centers,
        "total_rows": total_rows,
        "total_payable": total_payable,
        "ready_count": ready_count,
    }


def send_vendor_workbook_emails(vendor_files, sender_email, gmail_app_password, test_mode, test_email, report_name):
    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender_email, gmail_app_password)
    sent_count = 0
    send_log = []

    try:
        for item in vendor_files:
            if not test_mode and not item["Email"]:
                send_log.append({**item, "EmailStatus": "SKIPPED - Missing email"})
                continue

            if test_mode and not test_email.strip():
                send_log.append({**item, "EmailStatus": "SKIPPED - Missing test email"})
                continue

            to_email = test_email.strip() if test_mode else item["Email"]
            msg = EmailMessage()
            msg["Subject"] = build_email_subject(item["CenterName"], item["Identifier"], report_name)
            msg["From"] = sender_email
            msg["To"] = to_email
            msg["CC"] = item["CC"]
            msg.set_content(build_email_body(item["CenterName"], item["Identifier"], item["TotalRows"], item["PayableY"], report_name))
            msg.add_attachment(
                item["FileBytes"],
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=item["FileName"],
            )
            server.send_message(msg)
            sent_count += 1
            send_log.append({**item, "EmailStatus": f"SENT to {to_email}"})
    finally:
        server.quit()

    clean_log = []
    for row in send_log:
        clean_log.append({k: v for k, v in row.items() if k not in ["FileBytes"]})
    return sent_count, pd.DataFrame(clean_log)


def render_report_page(config):
    report_key = config["report_key"]
    report_name = config["report_name"]
    file_type = config["file_type"]

    back_col, title_col = st.columns([1, 6])
    with back_col:
        if st.button("← Back", key=f"back_{report_key}", width="stretch"):
            go_to("dashboard")
    with title_col:
        st.markdown(
            f"""
            <div style="font-size: 28px; font-weight: 800; color: #111827;">
                {report_name}
            </div>
            <div style="font-size: 15px; color: #6b7280; margin-top: 4px; margin-bottom: 16px;">
                Generate a review package, preview results, approve, then send center performance workbooks.
            </div>
            """,
            unsafe_allow_html=True
        )

    page_tab1, page_tab2 = st.tabs(["Run Report", "Center Profiles"])

    with page_tab2:
        profiles_df = load_center_profiles()
        cols_to_show = ["CenterName", "TeamEmail", config["profile_identifier_col"], "ContactPerson", "Notes", "Active"]
        cols_to_show = [c for c in cols_to_show if c in profiles_df.columns]
        st.subheader(f"{report_name} Profile Mapping View")
        st.dataframe(profiles_df[cols_to_show], width="stretch")
        st.caption(f"This report uses `{config['profile_identifier_col']}` from Center Profiles as its mapping source.")
        if st.button("Open Full Center Profiles", key=f"profiles_jump_{report_key}", width="stretch"):
            go_to("profiles")

    with page_tab1:
        controls_left, controls_mid, controls_right = st.columns([1, 1, 2])
        with controls_left:
            test_mode = st.checkbox("Test Mode", value=True, key=f"{report_key}_test_mode")
        with controls_mid:
            test_email = st.text_input("Test Email", key=f"{report_key}_test_email")
        with controls_right:
            st.info(f"Every email from this workflow is automatically CC'd to: {CC_EMAIL}")

        show_mode_banner(test_mode)
        render_upload_instructions(report_name, file_type)
        render_excel_tab_requirements(report_name)

        uploaded_file = st.file_uploader(
            f"Upload {report_name} file",
            type=["xlsx", "xls"] if file_type == "excel" else ["csv"],
            key=f"{report_key}_uploader"
        )

        package_key = f"{report_key}_report_package"
        approval_key = f"{report_key}_admin_approved"
        signature_key = f"{report_key}_upload_signature"

        if uploaded_file is not None:
            signature = f"{uploaded_file.name}_{uploaded_file.size}"
            if st.session_state.get(signature_key) != signature:
                st.session_state[signature_key] = signature
                st.session_state[package_key] = None
                st.session_state[approval_key] = False

        if uploaded_file is None:
            st.info(f"Upload a file to begin the {report_name.lower()} workflow.")
            return

        if not validate_uploaded_file(uploaded_file, file_type, report_name):
            return

        if st.button("Generate Review Package", key=f"{report_key}_generate_package", type="primary", width="stretch"):
            try:
                package = process_report_package(uploaded_file, config, test_mode, test_email)
                st.session_state[package_key] = package
                st.session_state[approval_key] = False
                st.success("Review package generated. Emails are locked until admin approval.")
            except Exception as e:
                st.session_state[package_key] = None
                st.session_state[approval_key] = False
                st.error(f"Something went wrong while processing this file: {e}")
                st.info("Please verify that you uploaded the correct report file and that the expected columns are present.")
                return

        package = st.session_state.get(package_key)
        if not package:
            return

        total_centers = package["total_centers"]
        total_rows = package["total_rows"]
        total_payable = package["total_payable"]
        ready_count = package["ready_count"]
        conversion_rate = f"{(total_payable / total_rows * 100):.1f}%" if total_rows > 0 else "0.0%"

        m1, m2, m3, m4, m5 = st.columns(5)
        with m1:
            metric_card("Centers", total_centers)
        with m2:
            metric_card("Total Leads", total_rows)
        with m3:
            metric_card("Payable Leads", total_payable)
        with m4:
            metric_card("Conversion Rate", conversion_rate)
        with m5:
            metric_card("Ready to Send", ready_count)

        tab1, tab2, tab3, tab4, tab5 = st.tabs(["Admin Review", "Downloads", "Email Queue", "Raw Preview", "Exceptions"])

        with tab1:
            st.subheader("Admin Review Required")
            st.warning("Emails are locked until the review workbook and on-screen previews are approved.")
            st.dataframe(package["summary_df"], width="stretch")
            if not package["disposition_summary"].empty:
                st.subheader("Disposition Summary")
                st.dataframe(package["disposition_summary"], width="stretch")

            reviewed_preview = st.checkbox("I reviewed the on-screen summary and email queue.", key=f"{report_key}_reviewed_preview")
            reviewed_workbook = st.checkbox("I downloaded/reviewed the admin workbook or ZIP package.", key=f"{report_key}_reviewed_workbook")
            approve_disabled = not (reviewed_preview and reviewed_workbook)

            if st.button("Approve Batch for Sending", key=f"{report_key}_approve_batch", type="primary", width="stretch", disabled=approve_disabled):
                st.session_state[approval_key] = True
                st.success("Batch approved. Email sending is now unlocked.")

            if st.button("Reset Approval", key=f"{report_key}_reset_approval", width="stretch"):
                st.session_state[approval_key] = False
                st.warning("Approval reset. Emails are locked again.")

            if st.session_state.get(approval_key):
                st.success("ADMIN APPROVAL COMPLETE - email sending is unlocked.")
            else:
                st.error("ADMIN APPROVAL REQUIRED - email sending is locked.")

        with tab2:
            st.subheader("Downloads")
            st.download_button(
                label="Download Admin Review Workbook",
                data=package["admin_workbook_bytes"],
                file_name=f"{report_key}_admin_review_workbook.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch"
            )
            st.download_button(
                label="Download ZIP of all center workbooks",
                data=package["zip_bytes"],
                file_name=f"{report_key}_center_workbooks.zip",
                mime="application/zip",
                width="stretch"
            )
            st.download_button(
                label="Download center summary CSV",
                data=package["summary_df"].to_csv(index=False).encode("utf-8"),
                file_name=f"{report_key}_center_summary.csv",
                mime="text/csv",
                width="stretch"
            )

        with tab3:
            st.subheader("Email Queue")
            st.dataframe(package["email_preview_df"], width="stretch")

            live_confirm = False
            if not test_mode:
                live_confirm = st.checkbox(
                    "I confirm I want to send emails to live vendor addresses.",
                    key=f"{report_key}_live_confirm"
                )

            send_disabled = False
            if not st.session_state.get(approval_key, False):
                send_disabled = True
                st.warning("Admin approval is required before emails can be sent.")
            if test_mode and not test_email.strip():
                send_disabled = True
                st.info("Enter a test email to enable sending in Test Mode.")
            if not test_mode and not live_confirm:
                send_disabled = True
                st.warning("Live sending requires confirmation before emails can be sent.")

            send_label = "Send TEST Emails" if test_mode else "Send LIVE Emails"
            if st.button(send_label, key=f"{report_key}_send_emails", type="primary", width="stretch", disabled=send_disabled):
                try:
                    sent_count, send_log_df = send_vendor_workbook_emails(
                        vendor_files=package["vendor_files"],
                        sender_email=sender_email,
                        gmail_app_password=gmail_app_password,
                        test_mode=test_mode,
                        test_email=test_email,
                        report_name=report_name,
                    )
                    st.success(f"Sent {sent_count} emails successfully.")
                    st.dataframe(send_log_df, width="stretch")
                    mode_value = "TEST" if test_mode else "LIVE"
                    log_report_run(report_name, total_rows, total_payable, total_centers, sent_count, mode_value, len(package["missing_email_df"]))
                    log_center_runs(report_name, package["summary_df"], mode_value)
                    if len(package["missing_email_df"]) > 0:
                        log_error_event(report_name, "Missing Emails", f"{len(package['missing_email_df'])} centers missing emails", len(package["missing_email_df"]), mode_value)
                    st.session_state[approval_key] = False
                except Exception as e:
                    try:
                        log_error_event(report_name, "Email Send Failure", str(e), 1, "TEST" if test_mode else "LIVE")
                    except Exception:
                        pass
                    st.error(f"Email error: {e}")

        with tab4:
            st.subheader("Raw Preview")
            st.dataframe(package["raw_df"].head(50), width="stretch")

        with tab5:
            st.subheader("Exceptions")
            if package["exception_df"].empty:
                st.success("No exceptions found.")
            else:
                st.dataframe(package["exception_df"], width="stretch")
            if not package["duplicate_df"].empty:
                st.subheader("Duplicate Leads")
                st.dataframe(package["duplicate_df"].head(200), width="stretch")

        st.info(f"This report maps `{config['identifier_label']}` to `{config['profile_identifier_col']}` in Center Profiles.")

# =========================
# SESSION + SECRETS
# =========================
sender_email = st.secrets["EMAIL"]
gmail_app_password = st.secrets["PASSWORD"]
site_password = st.secrets["APP_PASSWORD"]

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if "current_page" not in st.session_state:
    st.session_state.current_page = "dashboard"

if "selected_center_name" not in st.session_state:
    st.session_state.selected_center_name = ""


# =========================
# LOGIN SCREEN
# =========================
if not st.session_state.authenticated:
    st.markdown(
        """
        <style>
        .stApp {
            background: linear-gradient(180deg, #020617 0%, #0f172a 100%);
        }
        </style>
        """,
        unsafe_allow_html=True
    )

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown(
            """
            <div style="text-align: center; padding-top: 60px; padding-bottom: 20px;">
            """,
            unsafe_allow_html=True,
        )

        st.image(LOGO_FILE, width=260)

        st.markdown(
            """
            <div style="font-size: 30px; font-weight: 800; color: white; margin-top: 12px;">
                LIVMED Report Portal
            </div>
            <div style="font-size: 15px; color: #cbd5e1; margin-top: 8px; margin-bottom: 20px;">
                Secure access required
            </div>
            """,
            unsafe_allow_html=True,
        )

        entered_password = st.text_input("Access Password", type="password")
        login_clicked = st.button("Login", width="stretch")

        if login_clicked:
            if entered_password == site_password:
                st.session_state.authenticated = True
                st.session_state.current_page = "dashboard"
                st.rerun()
            else:
                st.error("Incorrect password.")

        st.markdown("</div>", unsafe_allow_html=True)

    st.stop()


# =========================
# MAIN HEADER
# =========================
show_top_header()

top_left, top_mid, top_right = st.columns([5, 1, 1])

with top_mid:
    if st.button("Dashboard", width="stretch"):
        st.session_state.current_page = "dashboard"
        st.rerun()

with top_right:
    if st.button("Log out", width="stretch"):
        st.session_state.authenticated = False
        st.session_state.current_page = "dashboard"
        st.rerun()


# =========================
# ROUTER
# =========================
current_page = st.session_state.current_page

if current_page == "dashboard":
    render_dashboard()

elif current_page == "profiles":
    render_center_profiles_page()

elif current_page == "profile_detail":
    render_center_profile_detail_page()

elif current_page in REPORT_CONFIGS:
    render_report_page(REPORT_CONFIGS[current_page])
